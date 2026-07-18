#!/usr/bin/env python3
"""
build_pad.py — Generador del pad.json de Vaca Viewer a partir de archivos crudos.

Copyright (c) 2026 Gonzalo Carvallo. Todos los derechos reservados.

QUÉ HACE
--------
Lee los archivos operativos crudos (surveys XLS, tallys PDF de OpenWells, fracplans XLS)
y produce el `pad.json` (schema_version 1) que la app Vaca Viewer sabe cargar.

Hoy este parseo corre FUERA de la app (este script Python). En v0.3 se migra a JS
dentro de la app (SheetJS para XLS, pdf.js para PDF) replicando exactamente esta lógica.

DEPENDENCIAS
------------
    pip install openpyxl pdfplumber --break-system-packages

USO
---
1. Poné los archivos crudos en el mismo directorio que este script (o ajustá las rutas
   en los diccionarios SURVEYS / FRACPLANS / TALLYS de más abajo).
2. Ajustá esos diccionarios con los nombres de archivo reales de tu pad y la metadata
   del pad (id, campaña, etc.) en el bloque `pad={...}`.
3. Corré:  python3 build_pad.py
4. Sale el JSON (por defecto a /tmp/pad_<ID>.json; cambialo si querés). Copialo a
   samples/ del repo o cargalo desde la sección Datos de la app.

CÓMO ADAPTARLO A OTRO PAD
-------------------------
- SURVEYS: {well_id: "archivo_survey.xlsx", ...}
- FRACPLANS: {well_id: "archivo_fracplan.xlsx", ...}
- TALLYS: {well_id: {"guia":..., "intermedia1":..., "intermedia2":..., "produccion":...}, ...}
  (podés omitir fases que no tengas; la app dibuja lo que exista)
- El bloque `pad={...}` al final: id, name, field, operator, campaign, wellhead_spacing_m.
- wellhead x = i*10 (bocas cada 10 m O→E). Ajustá si el pad tiene otra separación.

GOTCHAS IMPORTANTES (ya resueltos acá, NO re-romper)
----------------------------------------------------
1. SURVEYS tienen DOS layouts de columnas distintos. Por eso parse_survey detecta
   columnas por HEADER (MD/NS/EW/TVD/...), no por posición fija. Algunos surveys traen
   NS/EW como "S 0.05"/"W 0.02" (con letra), otros numéricos con signo. Y algunos traen
   además Northing/Easting ABSOLUTOS del CRS (~5.7 millones) que NO son los offsets
   locales — si se confunden, los pozos se dibujan a millones de metros (pantalla negra).
2. TALLYS: el OD del casing es el valor x.xxx (4.0–13.5) MÁS FRECUENTE en la sección
   "2.1 Pipe Sections". Esto descarta accesorios como "XCD 16"" (zapato perforador) que
   aparece en las guías y no es la cañería (que es 13⅜").
3. GRADO de acero de la aislación 5": a veces no está en el tally (usan designación de
   rosca W461/W463), entonces grade=None. Es correcto, no es un bug.
4. CAÑOS CORTOS (pup joints) de la aislación: se leen de "2.2 Run Tally". Regla del campo:
   caño NORMAL >10 m, caño corto <10 m. Su fila NO trae "N° Tally" (columna vacía), por eso el
   parser ubica los 3 decimales consecutivos (Longitud, Cum.Length, Set Depth) sin asumir cuántos
   enteros van antes. Set Depth = MD del TOPE del caño (≈ TD − Cum.Length) ⇒ bottom = Set Depth +
   Longitud. Si tu tally usa otra convención de Set Depth, ajustá parse_short_joints.
"""
import openpyxl, pdfplumber, re, json, warnings
from collections import Counter
from datetime import datetime, timezone
warnings.filterwarnings("ignore")

# ODs estándar de casing (pulg) y OD esperado por fase — para "encajar" candidatos y descartar ruido
STD_OD=[4.5,5.0,5.5,7.0,7.625,8.625,9.625,10.75,11.75,13.375]
PHASE_OD_EXP={"guia":13.375,"intermedia1":9.625,"intermedia2":7.625,"produccion":5.0}

def parse_survey(path):
    wb=openpyxl.load_workbook(path, data_only=True); ws=wb[wb.sheetnames[0]]
    hdr_row=None
    for r in range(1,40):
        v=ws.cell(r,2).value
        if isinstance(v,str) and v.strip().upper().startswith("MD"): hdr_row=r; break
    if hdr_row is None: hdr_row=23
    col={}
    for c in range(1, ws.max_column+1):
        h=ws.cell(hdr_row,c).value
        if isinstance(h,str):
            col[h.strip().upper().split("\n")[0].strip()]=c
    def find(*names):
        for n in names:
            if n in col: return col[n]
        return None
    c_md=find("MD"); c_incl=find("INCL"); c_az=find("AZIM GRID","AZIM","AZIMUTH")
    c_tvd=find("TVD"); c_vsec=find("VSEC"); c_ns=find("NS"); c_ew=find("EW"); c_dls=find("DLS")
    def num(x):
        if x is None: return None
        if isinstance(x,(int,float)): return float(x)
        m=re.search(r'-?[\d.]+', str(x).replace(",",""))
        return float(m.group()) if m else None
    def signed(x):
        if x is None: return None
        if isinstance(x,(int,float)): return float(x)
        m=re.search(r'([NSEW])\s*([\d.]+)', str(x).strip())
        if m:
            v=float(m.group(2)); return -v if m.group(1) in ("S","W") else v
        return num(x)
    vsec_az=None
    for r in range(5,22):
        for lc,vc in ((1,3),(7,9)):
            lab=ws.cell(r,lc).value
            if isinstance(lab,str) and "Vertical Section Azimuth" in lab:
                m=re.search(r'([\d.]+)', str(ws.cell(r,vc).value)); vsec_az=float(m.group(1)) if m else None
    stations=[]
    for r in range(hdr_row+1, ws.max_row+1):
        md=ws.cell(r,c_md).value if c_md else None
        if not isinstance(md,(int,float)): continue
        stations.append({
            "md":round(float(md),3),
            "incl":num(ws.cell(r,c_incl).value) if c_incl else None,
            "azim":num(ws.cell(r,c_az).value) if c_az else None,
            "tvd":round(num(ws.cell(r,c_tvd).value or 0),3) if c_tvd else None,
            "vsec":num(ws.cell(r,c_vsec).value) if c_vsec else None,
            "ns":signed(ws.cell(r,c_ns).value) if c_ns else None,
            "ew":signed(ws.cell(r,c_ew).value) if c_ew else None,
            "dls":num(ws.cell(r,c_dls).value) if c_dls else None,
        })
    return {"vsec_azimuth_deg":vsec_az, "stations":stations}

def parse_fracplan(path):
    wb=openpyxl.load_workbook(path, data_only=True); ws=wb["Punzados"]
    c=lambda r,col: ws.cell(r,col).value
    hdr={"lp_md":c(5,2),"collar_md":c(6,2),"horizontal_ext_m":c(7,2),"total_stages":c(9,2)}
    stages={}; order=[]; cur=None
    for r in range(13, ws.max_row+1):
        name=c(r,1)
        if not (isinstance(name,str) and name.startswith("Cluster")): continue
        n=int(re.search(r'(\d+)', name).group(1))
        stg=c(r,5); plug=c(r,12)
        if stg is not None:
            cur=int(stg)
            if cur not in stages: stages[cur]={"stage":cur,"plug_md":None,"clusters":[]}; order.append(cur)
        if cur is None: continue
        stages[cur]["clusters"].append({"n":n,"top_md":c(r,2),"bottom_md":c(r,3),
            "incl":c(r,4),"shots":c(r,8),"charge":c(r,9),"phasing":c(r,10)})
        if plug is not None: stages[cur]["plug_md"]=plug
    return {"total_stages":int(hdr["total_stages"]) if hdr["total_stages"] else len(order),
            "lp_md":hdr["lp_md"],"collar_md":hdr["collar_md"],
            "horizontal_ext_m":hdr["horizontal_ext_m"],"planned_vs_actual":"planned",
            "stages":[stages[s] for s in order]}

def parse_run_tally(text, max_len=10.0, edge_m=150):
    """Del '2.2 Run Tally' separa las piezas cortas (<10 m; los caños normales miden >10 m):
      - shorts: casing corto / XOVER del TRAMO MEDIO (se ignoran primeros/últimos `edge_m` m:
        cabezal/colgador arriba, shoetrack abajo).
      - shoetrack: piezas cortas de los últimos `edge_m` m (zapato, collar flotador, camisas, cortos).
    Cada fila trae 3 decimales consecutivos = Longitud, Cum.Length, Set Depth. La fila corta NO trae
    'N° Tally', por eso ubicamos los 3 decimales sin asumir cuántos enteros van antes. Set Depth = MD
    del TOPE (≈ TD − Cum.Length) ⇒ bottom = Set Depth + Longitud.
    """
    m=re.search(r'2\.2\s+Run\s+Tally(.*?)(?:\n\s*2\.3|\n\s*3\.|\Z)', text, re.S|re.I)
    block=m.group(1) if m else text
    dec=r'\d+(?:,\d{3})*\.\d+'
    row_re=re.compile(rf'({dec})\s+({dec})\s+({dec})')
    rows=[]
    for line in block.splitlines():
        if not re.match(r'\s*\d', line): continue        # fila de tally: empieza con el N° de junta
        mm=row_re.search(line)
        if not mm: continue
        length=float(mm.group(1).replace(",",""))
        setd=float(mm.group(3).replace(",",""))
        if length<=0 or setd<=0 or length>30: continue
        # tras el Set Depth pueden venir MÁS columnas numéricas (peso acumulado, etc.):
        # se descartan los decimales sueltos del principio de la "descripción"
        desc=re.sub(rf'^(?:{dec}\s+)+', '', line[mm.end():].strip())
        rows.append({"length":length, "setd":setd, "desc":desc})
    if not rows: return [], None
    max_md=max(r["setd"]+r["length"] for r in rows)      # fondo de la sarta (≈ zapato)
    shorts=[]; st_els=[]
    for r in rows:
        if r["length"]>=max_len: continue                # caño normal
        top=round(r["setd"],3); bottom=round(r["setd"]+r["length"],3)
        if top<edge_m: continue                          # primeros 150 m (cabezal/colgador): se ignora
        if bottom>max_md-edge_m:                          # últimos 150 m: shoetrack
            st_els.append({"desc":r["desc"] or "elemento", "length_m":round(r["length"],3),
                           "top_md":top, "bottom_md":bottom})
            continue
        is_xo=bool(re.search(r'\bXO\b|X-?OVER|CROSS', r["desc"], re.I))
        shorts.append({"desc":r["desc"] or "caño corto", "xover":is_xo,
                       "length_m":round(r["length"],3), "top_md":top, "bottom_md":bottom})
    shorts.sort(key=lambda x:x["top_md"])
    st_els.sort(key=lambda x:x["top_md"])
    return shorts, ({"elements":st_els} if st_els else None)

def parse_tally(path, phase=None):
    with pdfplumber.open(path) as pdf:
        text="\n".join((pg.extract_text() or "") for pg in pdf.pages)
    m=re.search(r'2\.1\s+Pipe\s+Sections(.*?)2\.2\s+Run\s+Tally', text, re.S|re.I)
    block=m.group(1) if m else text[:4000]
    # OD: encaja cada candidato al OD estándar más cercano (±0.07") para descartar longitudes
    # de tiro (~12 m); si se conoce la fase, prioriza su OD esperado.
    snapped=[]
    for x in re.findall(r'\b(\d{1,2}\.\d{1,3})\b', text):
        v=float(x); best=None; bd=0.07
        for s in STD_OD:
            d=abs(v-s)
            if d<bd: bd=d; best=s
        if best is not None: snapped.append(best)
    if phase and PHASE_OD_EXP.get(phase) is not None:
        snapped=[v for v in snapped if abs(v-PHASE_OD_EXP[phase])<1.6]
    od=Counter(snapped).most_common(1)[0][0] if snapped else None
    wts=[float(x) for x in re.findall(r'\b(\d{2,3}\.\d{1,2})\b', block)]
    wts=[w for w in wts if 14<=w<=120]
    weight=Counter(wts).most_common(1)[0][0] if wts else None
    txt2=block + text[:3000]
    grades=re.findall(r'\b([KNPL])[\s-]?(\d{2,3})Q?\b', txt2)
    norm=[a+b for a,b in grades if 50<=int(b)<=140]
    grade=Counter(norm).most_common(1)[0][0] if norm else None
    mt=re.search(r'Total\s+length\s+run\s+([\d,]+\.\d+)', text, re.I)
    shoe=float(mt.group(1).replace(",","")) if mt else None
    # caños cortos + shoetrack: solo interesan en la aislación (5"); en las otras fases se omiten
    shorts, shoetrack = parse_run_tally(text) if phase=="produccion" else ([], None)
    return od, shoe, weight, grade, shorts, shoetrack

# ============================================================================
# CONFIGURACIÓN DEL PAD — ajustar estos diccionarios y el bloque pad={} abajo.
# Nombres de archivo del pad MMo-35 (ejemplo). Reemplazar para otro pad.
# ============================================================================
SURVEYS={"MMo-2351h":"MMo-2351_h__Survey_Report_6668m_TD__1_.xlsx",
 "MMo-2352h":"PET_Nq_MMo-2352_h__Surveys_6633m_TD.xlsx",
 "MMo-2353h":"Standard_Survey_Report_MMo-2353_h___6628mTD.xlsx",
 "MMo-2354h":"Survey_Report_MMo-2354_h__6656m_TD.xlsx"}
FRACPLANS={w:f"{w}_Frac_Plan_29052026.xlsx" for w in SURVEYS}
TALLYS={
 "MMo-2351h":{"guia":"Drillers_Running_Tally_Report_2351_guia.pdf","intermedia1":"Drillers_Running_Tally_Report_2351_Int1.pdf","intermedia2":"Drillers_Running_Tally_Report_2351_Int2.pdf","produccion":"MMo-2351_-_Drillers_Running_Tally_Report_aislacion_5_pulg.pdf"},
 "MMo-2352h":{"guia":"Drillers_Running_Tally_Report_2352_guia.pdf","intermedia1":"Drillers_Running_Tally_Report_2352_Int1.pdf","intermedia2":"Drillers_Running_Tally_Report_2352_Int2.pdf","produccion":"MMo-2352_-_Drillers_Running_Tally_Report_Aislacion_5__MMo_2352__h_.pdf"},
 "MMo-2353h":{"guia":"Drillers_Running_Tally_Report_2353_guia.pdf","intermedia1":"Drillers_Running_Tally_Report_2353_Int1.pdf","intermedia2":"Drillers_Running_Tally_Report_2353_Int2.pdf","produccion":"MMo-2353_-_Drillers_Running_Tally_Report_-_Aislacion_-_MMO-2353h.pdf"},
 "MMo-2354h":{"guia":"Drillers_Running_Tally_Report_2354_guia.pdf","intermedia1":"Drillers_Running_Tally_Report_2354_Int1.pdf","intermedia2":"Drillers_Running_Tally_Report_2354_Int2.pdf","produccion":"MMo-2354_-__Tally_AISLACION_run_-_Centralizado.pdf"}}

WELL_ORDER=["MMo-2351h","MMo-2352h","MMo-2353h","MMo-2354h"]   # orden O→E (boca x = i*10)
PAD_ID="MMo-35"; PAD_NAME="Mata Mora PAD 35"; FIELD="ARC Phoenix - Mata Mora"
OPERATOR="Phoenix"; CAMPAIGN=2; SPACING=10; RKB=420.85
OUT=f"/tmp/pad_{PAD_ID}.json"
# ============================================================================

wells=[]
for i,wid in enumerate(WELL_ORDER):
    survey=parse_survey(SURVEYS[wid]); frac=parse_fracplan(FRACPLANS[wid])
    casings=[]
    for phase in ["guia","intermedia1","intermedia2","produccion"]:
        if phase in TALLYS.get(wid,{}):
            od,shoe,weight,grade,shorts,shoetrack=parse_tally(TALLYS[wid][phase], phase)
            cas={"phase":phase,"od_in":od,"shoe_md":shoe,"toc_md":None,"weight_ppf":weight,"grade":grade}
            if shorts: cas["short_joints"]=shorts
            if shoetrack: cas["shoetrack"]=shoetrack
            casings.append(cas)
    wells.append({"id":wid,"name":f"PET.Nq.{wid[:8]}(h)","well_number":int(wid[6]),
        "architecture":"horizontal",
        "wellhead":{"x":i*float(SPACING),"y":0.0,"rkb_elev_m":RKB,"ground_elev_m":RKB},
        "survey":survey,"casings":casings,"frac":frac})

pad={"schema_version":1,"app":"vaca-viewer",
  "generated_at":datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
  "pad":{"id":PAD_ID,"name":PAD_NAME,"field":FIELD,"operator":OPERATOR,"campaign":CAMPAIGN,
    "surface":{"wellhead_spacing_m":SPACING,"line_azimuth_deg":90},"wells":wells}}
json.dump(pad, open(OUT,"w",encoding="utf-8"), ensure_ascii=False, separators=(",",":"))

import os
print(f"=== PAD {PAD_ID} generado ===")
for w in wells:
    ew=[s["ew"] for s in w["survey"]["stations"]]; ns=[s["ns"] for s in w["survey"]["stations"]]
    cs=" · ".join(f'{c["phase"][:4]} {c["od_in"]}"@{round(c["shoe_md"])}' for c in w["casings"])
    nsj=sum(len(c.get("short_joints",[])) for c in w["casings"])
    print(f'{w["id"]}: EW[{min(ew):.0f},{max(ew):.0f}] NS[{min(ns):.0f},{max(ns):.0f}] {w["frac"]["total_stages"]}et | {cs} | caños cortos: {nsj}')
print(f"JSON: {OUT}  ({os.path.getsize(OUT)} bytes)")
